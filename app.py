import os
from collections import defaultdict
from datetime import date
from flask import Flask, render_template, request, redirect, url_for, Response
from models import db, Lancamento

app = Flask(__name__)

# ─────────────────────────────────────────────────────────────
# BANCO DE DADOS
# ─────────────────────────────────────────────────────────────
database_url = os.environ.get("DATABASE_URL", "sqlite:///database.db")
if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = database_url

db.init_app(app)

with app.app_context():
    db.create_all()


# ── Filtro Jinja2: formata float → R$ 1.000,00 ──
@app.template_filter("brl")
def brl_filter(value):
    try:
        value = float(value)
        neg = value < 0
        abs_val = abs(value)
        inteiro, decimal = f"{abs_val:.2f}".split(".")
        inteiro = "{:,}".format(int(inteiro)).replace(",", ".")
        resultado = f"R$ {inteiro},{decimal}"
        return f"-{resultado}" if neg else resultado
    except (TypeError, ValueError):
        return "R$ 0,00"


FORMAS_PAGAMENTO = [
    "PIX",
    "Boleto",
    "Dinheiro",
    "Elias - Santander Master",
    "Elias - Santander Visa",
    "Elias - Itaú Master",
    "Elias - Itaú Visa",
    "Elias - Conta Santander",
    "Elias - Conta Itaú",
    "Elias - Conta Bradesco",
    "Elias - iFood Beneficios"
]


def sort_competencias(competencias):
    """Ordena competências no formato MM/AAAA de forma crescente."""
    def key(c):
        try:
            m, a = c.split("/")
            return (int(a), int(m))
        except Exception:
            return (9999, 99)
    return sorted(competencias, key=key)


def competencia_atual():
    """Retorna a competência atual no formato MM/AAAA."""
    hoje = date.today()
    return f"{hoje.month:02d}/{hoje.year}"


@app.route("/")
def dashboard():
    todas_competencias = sort_competencias(
        set(r.competencia for r in Lancamento.query.all())
    )

    competencia_param = request.args.get("competencia", None)
    emprestado_sel  = request.args.get("emprestado", "exceto")

    if competencia_param is None:
        # Primeiro acesso: seleciona a competência atual automaticamente
        atual = competencia_atual()
        competencia_sel = atual if atual in todas_competencias else ""
    else:
        # Usuário escolheu explicitamente (incluindo "Todas" que passa string vazia)
        competencia_sel = competencia_param

    query = Lancamento.query
    if competencia_sel:
        query = query.filter_by(competencia=competencia_sel)

    todos = query.all()

    if emprestado_sel == "1":
        todos = [x for x in todos if "emprestado" in x.descricao.lower()]
    elif emprestado_sel != "todos":
        todos = [x for x in todos if "emprestado" not in x.descricao.lower()]

    receitas_list = [x for x in todos if x.tipo == "Receita"]
    despesas_list = [x for x in todos if x.tipo == "Despesa"]

    total_receitas = sum(x.valor for x in receitas_list)
    total_despesas = sum(x.valor for x in despesas_list)
    saldo = total_receitas - total_despesas

    todos_comp = Lancamento.query
    if competencia_sel:
        todos_comp = todos_comp.filter_by(competencia=competencia_sel)
    todos_comp = todos_comp.all()
    emp_list     = [x for x in todos_comp if "emprestado" in x.descricao.lower()]
    emp_receitas = sum(x.valor for x in emp_list if x.tipo == "Receita")
    emp_despesas = sum(x.valor for x in emp_list if x.tipo == "Despesa")
    emp_saldo    = emp_receitas - emp_despesas
    emp_count    = len(emp_list)

    if competencia_sel:
        labels_rec_desp = [competencia_sel]
        dados_receitas  = [total_receitas]
        dados_despesas  = [total_despesas]
    else:
        comp_set = sort_competencias(set(r.competencia for r in todos))
        rec_map  = defaultdict(float)
        desp_map = defaultdict(float)
        for x in todos:
            if x.tipo == "Receita":
                rec_map[x.competencia] += x.valor
            else:
                desp_map[x.competencia] += x.valor
        labels_rec_desp = comp_set
        dados_receitas  = [rec_map[c] for c in comp_set]
        dados_despesas  = [desp_map[c] for c in comp_set]

    rec_forma = defaultdict(float)
    for x in receitas_list:
        rec_forma[x.forma_pagamento] += x.valor
    rank_recebimento = sorted(rec_forma.items(), key=lambda t: t[1], reverse=True)

    desp_forma = defaultdict(float)
    for x in despesas_list:
        desp_forma[x.forma_pagamento] += x.valor
    rank_pagamento = sorted(desp_forma.items(), key=lambda t: t[1], reverse=True)

    return render_template(
        "dashboard.html",
        receitas=total_receitas,
        despesas=total_despesas,
        saldo=saldo,
        competencias=todas_competencias,
        competencia_sel=competencia_sel,
        emprestado_sel=emprestado_sel,
        emp_receitas=emp_receitas,
        emp_despesas=emp_despesas,
        emp_saldo=emp_saldo,
        emp_count=emp_count,
        labels_rec_desp=labels_rec_desp,
        dados_receitas=dados_receitas,
        dados_despesas=dados_despesas,
        rank_recebimento=rank_recebimento,
        rank_pagamento=rank_pagamento,
    )


@app.route("/lancamentos", methods=["GET", "POST"])
def lancamentos():
    if request.method == "POST":
        novo = Lancamento(
            tipo=request.form["tipo"],
            competencia=request.form["competencia"],
            descricao=request.form["descricao"],
            valor=float(request.form["valor"]),
            forma_pagamento=request.form["forma_pagamento"]
        )
        db.session.add(novo)
        db.session.commit()
        return redirect(url_for("relatorios"))

    prefill = {
        "tipo": request.args.get("tipo", ""),
        "competencia": request.args.get("competencia", ""),
        "descricao": request.args.get("descricao", ""),
        "valor": request.args.get("valor", ""),
        "forma_pagamento": request.args.get("forma_pagamento", ""),
    }
    replicando = any(prefill.values())
    return render_template("lancamentos.html", formas_pagamento=FORMAS_PAGAMENTO, prefill=prefill, replicando=replicando)


@app.route("/relatorios")
def relatorios():
    todas_competencias = sort_competencias(
        set(r.competencia for r in Lancamento.query.all())
    )

    # Usa None como sentinel: se o parâmetro não veio na URL, aplica o default
    # Se veio explicitamente (mesmo vazio, via botão "Todas"), respeita a escolha
    competencia_param = request.args.get("competencia", None)

    if competencia_param is None:
        # Primeiro acesso: mostra todos os lançamentos por padrão
        competencia_sel = ""
    else:
        # Usuário escolheu explicitamente (incluindo "Todas" que passa string vazia)
        competencia_sel = competencia_param

    query = Lancamento.query
    if competencia_sel:
        query = query.filter_by(competencia=competencia_sel)

    dados = query.order_by(Lancamento.id.desc()).all()

    return render_template(
        "relatorios.html",
        dados=dados,
        competencias=todas_competencias,
        competencia_sel=competencia_sel,
    )


@app.route("/relatorios/exportar")
def exportar_xlsx():
    try:
     import io
     import openpyxl
     from openpyxl.styles import Font, PatternFill, Alignment
     from openpyxl.utils import get_column_letter
    except ImportError:
        return "Erro: openpyxl nao instalado. Execute: pip install openpyxl", 500

    competencia_sel = request.args.get("competencia", "").strip()

    query = Lancamento.query
    if competencia_sel:
        query = query.filter_by(competencia=competencia_sel)
    dados = query.order_by(Lancamento.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lançamentos"

    # Cabeçalho
    headers = ["#", "Tipo", "Competência", "Descrição", "Valor (R$)", "Forma de Pagamento"]
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for row_idx, item in enumerate(dados, 2):
        valor = item.valor if item.tipo == "Receita" else -item.valor
        ws.cell(row=row_idx, column=1, value=item.id)
        ws.cell(row=row_idx, column=2, value=item.tipo)
        ws.cell(row=row_idx, column=3, value=item.competencia)
        ws.cell(row=row_idx, column=4, value=item.descricao)
        ws.cell(row=row_idx, column=5, value=valor)
        ws.cell(row=row_idx, column=6, value=item.forma_pagamento)

        # Cor por tipo
        if item.tipo == "Receita":
            ws.cell(row=row_idx, column=5).font = Font(color="16A34A")
        else:
            ws.cell(row=row_idx, column=5).font = Font(color="DC2626")

    # Largura das colunas
    larguras = [8, 12, 16, 40, 18, 30]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nome = f"lancamentos_{competencia_sel.replace('/', '-') if competencia_sel else 'todos'}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome}"}
    )


@app.route("/replicar/<int:id>")
def replicar(id):
    item = Lancamento.query.get_or_404(id)
    return redirect(url_for("lancamentos",
        tipo=item.tipo,
        competencia=item.competencia,
        descricao=item.descricao,
        valor=item.valor,
        forma_pagamento=item.forma_pagamento
    ))


@app.route("/editar/<int:id>", methods=["GET", "POST"])
def editar(id):
    item = Lancamento.query.get_or_404(id)

    if request.method == "POST":
        item.tipo            = request.form["tipo"]
        item.competencia     = request.form["competencia"]
        item.descricao       = request.form["descricao"]
        item.valor           = float(request.form["valor"])
        item.forma_pagamento = request.form["forma_pagamento"]
        db.session.commit()
        return redirect(url_for("relatorios"))

    return render_template("editar.html", item=item, formas_pagamento=FORMAS_PAGAMENTO)


@app.route("/excluir/<int:id>")
def excluir(id):
    item = Lancamento.query.get_or_404(id)
    db.session.delete(item)
    db.session.commit()
    return redirect(url_for("relatorios"))


if __name__ == "__main__":
    app.run(debug=True)
